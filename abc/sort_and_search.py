import pandas as pd
import numpy as np
from pandas import DataFrame
import openpyxl

# loading DataFrame from Excel files
df = pd.read_excel('canteen_db.xlsx')
infocan = pd.read_excel('canteen details.xlsx')
admin_data = pd.read_excel("Admin.xlsx")

# loading the worksheet from the excel file
wb = openpyxl.load_workbook('Canteen_db - Copy.xlsx')
ws = wb['Sheet1']

#  the name of the canteens to be the indexes of infocan and df
infocan = infocan.set_index('Canteen')
df = df.set_index(['Canteen'])

# takes in foodtype = ['Food1','food2' etc], pricerange = [lower, higher as floats/int], the search term
# rating = int(1 to 5) or 0 if not specified
def searchfood(foodtype, pricerange, rating, search, df):
    # copy the dataframe to temporary.
    search_df = df
    # filter by foodtype
    if foodtype != []:
        foodcond = search_df['Food Type'].isin(foodtype)
        search_df = search_df[foodcond]
    # filter by price range
    low_price = pricerange[0]
    high_price = pricerange[1]
    pricecond1 = search_df['Price'] >= low_price
    pricecond2 = search_df['Price'] <= high_price
    search_df = search_df[ pricecond1 & pricecond2 ]
    # filter by rating, shows all above specified rating
    if rating != 0:
        ratingcond = search_df['Rating'] >= rating
        search_df = search_df[ ratingcond ]
    # filter by menu Item
    if search != ' ':
        lst = search.lower().split()
        for word in lst:
            wordcond = search_df['Menu Item'].str.lower().str.contains(word)
            search_df = search_df[wordcond]
    # return the filtered DataFrame
    return search_df

# function to sort by rating given the DataFrame filtered, the output is a DataFrame
def sort_by_rating(filter_df):
    return filter_df.sort_values("Rating")

# function to sort by price given the DataFrame filtered, the output is a DataFrame
def sort_by_price(filter_df):
    return filter_df.sort_values("Price")

# function to sort by distance based on the user location, the filtered DataFrame and the DataFrame
# information about the canteen
def sort_by_location(user_loc, filter_df, infocan):
    lst_loc = filter_df.index.unique()
    lst_dist = {}
    frames = []
    for loc in lst_loc:
        # calculate the distance
        distance = ((user_loc[0] - infocan.loc[loc]['loc x'])**2 + (user_loc[1] - infocan.loc[loc]['loc y'])**2)**(1/2)
        # convert from bitmap to km (just giving an approximate of the distance)
        distance *= 0.0025
        # create a new column in DataFrame to store all of the distances
        filter_df.at[loc,'Distance'] = distance
    # sort DataFrame according to the distance
    filter_df = filter_df.sort_values('Distance')
    # select top 10 location to show the user
    count = 1
    for loc in filter_df.index.unique():
        a = filter_df.loc[loc]
        if len(a.shape) == 1:
            b = pd.DataFrame(a).T
            frames.append(b)
        else:
            frames.append(a)
        if count == 10:
            break
        count += 1
    if frames !=[]:
        #concatenate all the DataFrames selected
        return pd.concat(frames)
    else:
        return pd.DataFrame()

# display only 10 canteens satisfying the sort and search
def display10(filter_df):
    canteen_list_10 = filter_df.index.unique()[:10]
    return filter_df.loc[canteen_list_10,:]

# if filter_df has only one result for canteen X, its type of location would be np.fload64
#else it returns a series
def get_location(filter_df):
    df = filter_df
    if type(filter_df) == np.float64:
        return [df]
    else: return df

# searching the worksheet to find the row containing the canteen, stall and food
def search(ws, lst):
    rows = []
    for row_num in range (2, ws.max_row+1):
        can = ws.cell(row = row_num, column = 1).value
        stall = ws.cell(row = row_num, column = 3).value
        food = ws.cell(row = row_num, column = 4).value
        #if lst[2] == 'n' then we just need to assign a value to food
        #suc that food == lst[2]
        if lst[2] == 'n':
            food = 'n'
        if can == lst[0] and stall == lst[1] and food == lst[2]:
            rows.append(row_num)
    return rows
