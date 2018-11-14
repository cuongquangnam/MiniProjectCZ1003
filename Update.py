import openpyxl
import pandas as pd
from pandas import DataFrame
#load the worksheet from the excel file
wb = openpyxl.load_workbook('Canteen_db - Copy.xlsx')
ws = wb['Sheet1']
#function to search given list (canteen, stall, food), output the list of
#rows containing the information
def search(ws, lst):
    rows = []
    for row_num in range (2, ws.max_row+1):
        can = ws.cell(row = row_num, column = 1).value
        stall = ws.cell(row = row_num, column = 3).value
        food = ws.cell(row = row_num, column = 4).value
        #if lst[2] == 'n' then we just need to assign a value to food
        #suc that food == lst[2]
        if lst[2] == 'n':
            food = 'n'
        if can == lst[0] and stall == lst[1] and food == lst[2]:
            rows.append(row_num)
    return rows
#function to edit the worksheet
def edit(ws, row_num, foodtype, price, rating):
    lst = []
    for column_num in range(1,5):
        lst.append(ws.cell(column = column_num, row = row_num).value)
    lst[2] = foodtype
    lst.extend([price, rating])
    if price != ws.cell(column = 5, row = row_num).value:
        ws.delete_rows(row_num)
        if price <= ws.cell(column = 5, row = 2).value:
            row_num = 2
        elif price >= ws.cell(column = 5, row = ws.max_row).value:
            row_num = ws.max_row + 1
        else:
            row_num = binarySelecting(ws, price, 2, ws.max_row)
        ws.insert_rows(row_num)
    for i in range (1,7):
        _ = ws.cell(column = i, row = row_num, value = lst[i - 1])
#function to remove some information in the worksheet
def remove(ws):
    info = ask()
    result = search(ws, info)
    x = input('Are you sure? (y/n)')
    if x == 'y':
        if result == []:
            print('Oops! We can\'t find the food you inquire')
        else:
            result.reverse()
            for row_num in result:
                ws.delete_rows(row_num)
#function to add information into the worksheet
#location, foodtype, stall, menu_item, price, rating
def add(ws, lst):
    max = ws.max_row
    if lst[4] <= ws.cell(column = 5, row = 2).value:
        index = 2
    elif lst[4] >= ws.cell(column = 5, row = max).value:
        index = max + 1
    else:
        index = binarySelecting(ws, lst[4], 2, max)
    ws.insert_rows(index)
    for i in range (1,7):
        _ = ws.cell(column = i, row = index, value = lst[i - 1])
    #insert to the txt file
    with open("menu_items.txt", "a") as myfile:
        myfile.write(lst[3])
def binarySelecting(ws, target, low, high):
    if high - 1 == low:
        return high
    mid = (low + high)//2
    if target >= ws.cell(column = 5, row = mid).value and target <= ws.cell(column = 5, row = mid).value:
        return mid + 1
    elif target <  ws.cell(column = 5, row = mid).value:
        return binarySelecting(ws, target, low, mid)
    else:
        return binarySelecting(ws, target, mid, high)
#remove(ws)
#remember to save the workbook after updating the information
#wb.save('Canteen_db - Copy.xlsx')
#login()
#edit(ws, 15, 0,10)
add(ws,['Location','Foodtype', 'Stall','Cuong Food', 100, 3])
wb.save('Canteen_db - Copy.xlsx')
