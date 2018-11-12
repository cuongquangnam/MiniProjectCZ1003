import openpyxl
import pandas as pd
from pandas import DataFrame
#load the worksheet from the excel file
wb = openpyxl.load_workbook('Canteen_db - Copy.xlsx')
ws = wb['Sheet1']
#this is for the admin to login into their account
def login():
    df = pd.read_excel('Admin.xlsx')
    check = False
    while (check == False):
        admin = input('Admin??')
        password = input('Password??')
        condadmin = df['Admin account'] == admin
        condpass = df['Password'] == password
        result = df[condadmin & condpass]
        if result.empty:
            print('Invalide username or password')
        else:
            check = True
#this is a fucntion to ask the user basic information
def ask():
    canteen = input("Canteen??")
    stall = input('Stall??')
    #food == 'n' means that the user don't want to input the food name,
    #they just want to deal with the canteen and stall
    food = input('Food??')
    return (canteen, stall, food)
#function to search given list (canteen, stall, food), output the list of
#rows containing the information
def search(ws, lst):
    rows = []
    for row_num in range (2, ws.max_row+1):
        can = ws.cell(row = row_num, column = 1).value
        stall = ws.cell(row = row_num, column = 3).value
        food = ws.cell(row = row_num, column = 4).value
        #if lst[2] == 'n' then we just need to assign a value to food
        #suc that food == lst[2]
        if lst[2] == 'n':
            food = 'n'
        if can == lst[0] and stall == lst[1] and food == lst[2]:
            rows.append(row_num)
    return rows
#function to edit the worksheet
def edit(ws):
    info =ask()
    result = search(ws, info)
    if result == []:
        print('Oops! We can\'t find the food you inquire')
    else:
        #info[2] == 'n' means that they want to edit the rating of the stall
        #else they want to edit the price of the food
        if info[2] == 'n':
            rating = int(input('Rating'))
            x = input('Are you sure? (y/n)')
            if x == 'y':
                for row_num in result:
                    ws.cell(row = row_num, column = 6, value = rating)
        else:
            price = int(input('Price'))
            x = input('Are you sure? (y/n)')
            if x == 'y':
                for row_num in result:
                    ws.cell(row = row_num, column = 5, value = price)
#function to remove some information in the worksheet
def remove(ws):
    info = ask()
    result = search(ws, info)
    x = input('Are you sure? (y/n)')
    if x == 'y':
        if result == []:
            print('Oops! We can\'t find the food you inquire')
        else:
            result.reverse()
            for row_num in result:
                ws.delete_rows(row_num)
#function to add information into the worksheet
def add(ws):
    info = ask()
    result = search(ws, info)
    x = input('Are you sure? (y/n)')
    if x == 'y':
        if result != []:
            print("Sorry! The database have already had this information")
        else:
            price = input('Price??')
            rating = input('Rating??')
            foodtype = input('Type of food??')
            ws.append([info[0],foodtype, info[1], info[2], price, rating])
#remove(ws)
#remember to save the workbook after updating the information
add(ws)

wb.save('Canteen_db - Copy.xlsx')


